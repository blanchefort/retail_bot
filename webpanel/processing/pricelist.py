"""парсинг прайс-листов
"""
import os
import openpyxl
from django.conf import settings
from webpanel.models.product import Product
from webpanel.models.product_unit_type import ProductUnitType
from webpanel.models.product_category import ProductCategory

def get_data(price_file=''):
    """Парсит xlsx-файл и отдаёт их в виде словаря

    Аргументы:

    file (str) - путь к файлу для парсинга

    Возвращает:

    * сообщение об ошибке. Возвращаемые значения:
        true - всё нормально, 
        false - при парсинге возникла ошибка, данные не получены
        null - имеются пропуски в данных
    * данные в виде cписка словарей. Пример:

    [{'number': 1,
      'product_name': 'Knauf-Суперпол влагостойкий 20x600x1200 мм',
      'product_price': 367,
      'product_unit': 'м2'},
     {'number': 2,
      'product_name': 'Керамогранит «Forest» 20х60 см 1.08 м2 цвет медовый',
      'product_price': 868,
      'product_unit': 'м2'},
     {'number': 3,
      'product_name': 'Вагонка ПВХ 3000x100x10 мм, цвет белый, 0.3 м²',
      'product_price': 41.4,
      'product_unit': 'м2'}]
    """
    data = []
    missdata = False
    price_file = os.path.join(settings.MEDIA_ROOT, price_file)

    try:
        price_obj = openpyxl.load_workbook(price_file)
    except:
        return False, data

    sheet = price_obj.active

    if sheet.max_row < 4 or sheet.max_column < 4:
        return False, data

    
    for cell in range(4, sheet.max_row + 1):
        d = {}
        number = 1
        try:
            d['number'] = number
            number += 1
            point = 'B' + str(cell)
            d['product_name'] = str(sheet[point].value.strip())
            point = 'C' + str(cell)
            d['product_unit'] = str(sheet[point].value.strip())
            point = 'D' + str(cell)
            d['product_price'] = float(sheet[point].value)
            data.append(d)
            number += 1
        except:
            missdata = True
            break

    # посмотрим, нет ли пропусков
    for row in data:
        if len(row['product_name']) < 2 or row['product_price'] < 1:
            missdata = True
            break

    if missdata == True:
        return None, data

    return True, data


def save_products(price_file, user):
    """Сохранение товаров из прайс-листа в БД
    Аргументы:

    file (str) - путь к файлу для парсинга
    user (django.contrib.auth.models.User) - модель пользователя, к которому привязываются товары
    """
    _, products = get_data(price_file=price_file)

    # Сначала установим все товары данного пользователя неактивными
    Product.objects.filter(user=user).update(is_active=False)

    # теперь активируем или добавим те товары, который в новом прайс-листе
    for item in products:
        product_unit = ProductUnitType.objects.get_or_create(short=item['product_unit'])
        product_category = ProductCategory.objects.get_or_create(name='Без категории')

        if Product.objects.filter(user=user).filter(title=item['product_name']):
            p = Product.objects.filter(user=user).get(title=item['product_name'])
            p.price = item['product_price']
            p.is_active = True
            p.save()
        else:
            Product(
                title=item['product_name'],
                user=user,
                unit = product_unit[0],
                category = product_category[0],
                price = item['product_price'],
                is_active = True
            ).save()