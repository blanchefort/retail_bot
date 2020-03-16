import uuid
from datetime import datetime
from openpyxl import Workbook

from django.contrib.auth.decorators import login_required
from django.template.response import TemplateResponse
from django.core.exceptions import PermissionDenied
from django.http import HttpResponseNotFound
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from django.core.paginator import Paginator
from django.contrib.auth.models import User

from webpanel.forms import UploadFileForm
from webpanel.forms import UploadBillForm
from webpanel.forms import UpdateProfileSellerForm
from webpanel.processing.pricelist import get_data
from webpanel.processing.pricelist import save_products
from webpanel.processing.get_order import get_seller_order_by_number
from webpanel.models.price_list import PriceLists
from webpanel.models.product import Product
from webpanel.models.profile import Profile
from webpanel.models.order import Order
from webpanel.models.seller_bill import SellerBill

@login_required(login_url='/accounts/login/')
def index(request):
    """Точка входа профиля продавца
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Сводка по вашему аккаунту',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    # Распределение заказов
    # новые, 
    orders_new = Order.objects.filter(product__user=request.user).filter(status=1).values('user').distinct()
    context.update({'orders_new': len(orders_new)})
    # открытые, 
    orders_opened = Order.objects.filter(product__user=request.user).filter(status=2).values('order_number').distinct()
    context.update({'orders_opened': len(orders_opened)})
    # закрытые
    orders_closed = Order.objects.filter(product__user=request.user).filter(status=4).values('order_number').distinct()
    context.update({'orders_closed': len(orders_closed)})

    # Топ-5 товаров
    
    return TemplateResponse(request, 'seller/index.html', context=context)


@login_required(login_url='/accounts/login/')
def upload_price(request):
    """Загрузить прайс-лист
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Загрузить прайс-лист',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    if request.method == 'POST':
        # Проверяем одобрение прайс-листа
        approving = request.POST.get('approve_field')
        rejecting = request.POST.get('reject_field')
        price_id = request.POST.get('price_id')
        if approving == 'approved':
            # Файл одобрен, сохраняем в БД
            p = PriceLists.objects.get(id=price_id)
            save_products(p.file_name.name, request.user)
            messages.success(request, 'Товары успешно обновлены!')
            return redirect('seller_products')
        elif rejecting == 'rejected':
            # Файл не одобрен, удаляем данные
            PriceLists.objects.filter(id=price_id).delete()
            messages.info(request, 'Файл удалён.')
            upload_form = UploadFileForm()
        else:
            upload_form = UploadFileForm(request.POST, request.FILES)
            if upload_form.is_valid():
                uploaded_price = upload_form.save(commit=False)
                uploaded_price.user = request.user
                uploaded_price.save()
                messages.success(request, 'Файл успешно загружен!')

                # Выводим данные файла на проверку.
                error_state, product_list = get_data(price_file=uploaded_price.file_name.name)
                if error_state == False:
                    uploaded_price.delete()
                    messages.error(request, 'Данные не удалось загрузить. \
                        Проверьте, пожалуйста, корректен ли ваш файл и загрузите его ещё раз.')
                elif error_state == None:
                    uploaded_price.delete()
                    messages.error(request, 'Данные не сохранены. \
                        В таблице файла обнаружены пропуски. \
                        Проверьте, пожалуйста, корректен ли ваш файл и загрузите его ещё раз.')
                else:
                    context.update({'product_list': product_list})
                    context.update({'product_count': len(product_list)})
                    context.update({'product_link': uploaded_price.file_name.url})
                    context.update({'price_id': uploaded_price.id})
            else:
                messages.error(request, 'Загружен некорректный формат файла. Используйте xlsx-файлы.')

    else:
        upload_form = UploadFileForm()
        
    context.update({'upload_form': upload_form})
    return TemplateResponse(request, 'seller/upload_price.html', context=context)


@login_required(login_url='/accounts/login/')
def pricelists(request):
    """Список прайс-листов
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    lists = PriceLists.objects.filter(user=request.user).order_by('-uploaded')
    paginator = Paginator(lists, 20)
    page = request.GET.get('page')
    list = paginator.get_page(page)
    count = len(list)
    context = {
        # Список прайс-листов
        'title': 'Список прайс-листов',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username,
        'list': list,
        'count': count
    }
    return TemplateResponse(request, 'seller/pricelists.html', context=context)


@login_required(login_url='/accounts/login/')
def products(request):
    """Список товаров
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    products_all = Product.objects.filter(user=request.user).filter(is_active=True)
    paginator = Paginator(products_all, 20)
    page = request.GET.get('page')
    list = paginator.get_page(page)
    count = len(list)
    context = {
        # Заголовок текущей страницы
        'title': 'Список товаров',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username,
        'list': list,
        'count': count
    }
    return TemplateResponse(request, 'seller/products.html', context=context)


@login_required(login_url='/accounts/login/')
def orders(request):
    """Счета и заказы
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Счета и заказы
        'title': 'Счета и заказы',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    # Статус заявки:
    #         0 - не сформирован, 
    #         1 - отправлен продавцу
    #         2 - продавец обработал, выставил счёт
    #         3 - транспортник принял к исполнению
    #         4 - исполнен
    #         5 - удалён покупателем (удалить можно только с уровня 0, т.е. удаление из корзины,
    #         пока не был передан продавцу).
    # Новые заказы
    new_orders_count = Order.objects.filter(product__user=request.user
        ).filter(status=1
        ).count()
    # Заказы, ожидающие оплаты
    wait_orders_count = Order.objects.filter(product__user=request.user
        ).exclude(status=0
        ).exclude(status=1
        ).exclude(status=4
        ).exclude(status=5).count()
    # Выполненные заказы
    ready_orders_count = Order.objects.filter(product__user=request.user
        ).filter(status=4).count()
    # Всего заказов
    total_orders_count = Order.objects.filter(product__user=request.user
        ).exclude(status=0
        ).exclude(status=5).count()

    context.update({'new_orders_count': new_orders_count})
    context.update({'wait_orders_count': wait_orders_count})
    context.update({'ready_orders_count': ready_orders_count})
    context.update({'total_orders_count': total_orders_count})

    if new_orders_count == 0:
        messages.info(request, 'У вас нет новых заказов.')
    else:
        messages.success(request, f'У вас {new_orders_count} новых заказа.')
        new_order = Order.objects.filter(product__user=request.user
            ).filter(status=1
            )
        context.update({'new_order': new_order})

    # список счетов продавца
    bills = Order.objects.filter(product__user=request.user
        ).exclude(status=0
        ).exclude(status=5)
    if bills.count() > 0:
        # имеются счета, формируем список
        bill_set = set()
        bs = []
        for item in bills:
            if item.order_number not in bill_set:
                bill_set.update([item.order_number])
                current_bill = Order.objects.filter(order_number=item.order_number)
                current_bill_sum = 0
                for i in current_bill:
                    current_bill_sum += i.product_count * i.product.price
                b = {
                    'order_number': item.order_number,
                    'user': item.user.username,
                    'status': item.status,
                    'bill_sum': current_bill_sum,
                    'date': current_bill.first().creation_date
                }
                bs.append(b)

        #bs = bs.reverse()
        context.update({'bills': bs})
    return TemplateResponse(request, 'seller/orders.html', context=context)

@login_required(login_url='/accounts/login/')
def confirm_order(request, user_id, status):
    """Создание заказа для определённого покупателя
    """
    if request.user.profile.type != 4:
        raise PermissionDenied
    context = {
        # Заголовок текущей страницы
        'title': 'Создание заказа',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    # Изменяем статус заказа
    if status == 2:
        order = Order.objects.filter(product__user=request.user
            ).filter(status=1
            ).filter(user=user_id)
        if len(order) > 0:
            #order_number = uuid.uuid1()
            order_number=int(datetime.now().timestamp())
            Order.objects.filter(product__user=request.user
                ).filter(status=1
                ).filter(user=user_id).update(
                    status=2,
                    bill_date=datetime.now(),
                    # Python int too large to convert to SQLite INTEGER
                    #order_number=order_number.int
                    order_number=order_number
                )
            messages.info(request, 'Заказ принят в обработку. Сообщение об этом отправлено покупателю.')

    # Список заказа
    order = Order.objects.filter(product__user=request.user
        ).filter(status=1
        ).filter(user=user_id) or Order.objects.filter(product__user=request.user
        ).filter(status=2
        ).filter(user=user_id)
    if len(order) > 0:
        context.update({'order': order})
        context.update({'status': order.first().status})
        # Сумма заказа
        total_sum = 0
        for item in order:
            total_sum += item.product_count * item.product.price
        context.update({'total_sum': total_sum})

        # Покупатель
        order_user = User.objects.get(id=user_id)
        context.update({'order_user': order_user})
    else:
        raise HttpResponseNotFound


    return TemplateResponse(request, 'seller/confirm_order.html', context=context)

@login_required(login_url='/accounts/login/')
def order_details(request, order_number):
    """показывает детали подтверждённого заказа
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Детали заказа',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    order_items = Order.objects.filter(order_number=order_number)

    if order_items.count() > 0:
        context.update({'order_items': order_items})
        context.update({'order_number': order_number})
        context.update({'order_user': order_items.first().user.username})
        # Сумма заказа
        total_sum = 0
        for item in order_items:
            total_sum += item.product_count * item.product.price
        context.update({'order_sum': total_sum})
        context.update({'order_status': order_items.first().status})

        # Форма на отправку счёта
        if request.method == 'POST':
            print(1)
            upload_form = UploadBillForm(request.POST, request.FILES)
            print(2)
            if upload_form.is_valid():
                print(3)
                bill_price = upload_form.save(commit=False)
                bill_price.seller = request.user
                bill_price.user = order_items.first().user
                bill_price.order_number = order_number
                bill_price.order_sum = total_sum
                bill_price.reseived_flag = 0
                bill_price.save()
                messages.success(request, 'Файл успешно загружен!')
        else:
            context.update({'form': UploadBillForm()})

        # проверяем, есть ли выставленный счёт на этот заказ
        if SellerBill.objects.filter(order_number=order_number):
            context.update({'bill': SellerBill.objects.get(order_number=order_number)})

    else:
        raise HttpResponseNotFound

    return TemplateResponse(request, 'seller/order_details.html', context=context)

@login_required(login_url='/accounts/login/')
def download_order_as_xsls(request, user_id):
    """Скачать неподтверждённый заказ в виде эксель-файла
    """
    if request.user.profile.type != 4:
        raise PermissionDenied
    
    # Список заказа
    order = Order.objects.filter(product__user=request.user
        ).filter(status=1
        ).filter(user=user_id)
    
    if len(order) > 0:
        # Создаём эксель-таблицу
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=order-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        wb = Workbook()
        worksheet = wb.active
        worksheet.title = f'Заказ'

        columns = [
            '№',
            'Товар',
            'Цена',
            'Кол-во',
            'Ед.изм.',
            'Сумма'
        ]
        row_num = 1

        # Прописываем заголовки
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Заполняем ячейки данными
        counter = 1
        total_sum = 0 
        for item in order:
            product_sum = item.product_count * item.product.price
            row_num += 1
            row = [
                counter,
                item.product.title,
                item.product.price,
                item.product_count,
                item.product.unit.short,
                product_sum
            ]
            total_sum += product_sum
            counter += 1
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        row_num += 1
        cell = worksheet.cell(row=row_num, column=5)
        cell.value = 'Итого:'
        cell = worksheet.cell(row=row_num, column=6)
        cell.value = total_sum
        wb.save(response)
        return response
    else:
        raise HttpResponseNotFound

@login_required(login_url='/accounts/login/')
def customers(request):
    """Список покупателей
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Список покупателей',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }
    return TemplateResponse(request, 'seller/customers.html', context=context)


@login_required(login_url='/accounts/login/')
def requisites(request):
    """Реквизиты
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Реквизиты',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    if request.method == "POST":
        form = UpdateProfileSellerForm(data=request.POST, instance=request.user.profile)
        if form.is_valid():
            # Проверка
            phone = form.cleaned_data.get('phone')
            check_phone = Profile.objects.filter(phone=phone).exclude(user=request.user)

            if check_phone.count() == 0:
                bin = form.cleaned_data.get('bin')
                check_bin = Profile.objects.filter(bin=bin).exclude(user=request.user)

                if check_bin.count() == 0:
                    request.user.profile.phone = form.cleaned_data.get('phone')
                    request.user.profile.company_name = form.cleaned_data.get('company_name')
                    request.user.profile.address = form.cleaned_data.get('address')
                    request.user.profile.bin = form.cleaned_data.get('bin')
                    request.user.profile.bank_account = form.cleaned_data.get('bank_account')
                    request.user.save()
                    messages.success(request, 'Реквизиты вашей организации обновлены.')
                else:
                    messages.error(request, 'Данный БИН нельзя использовать.')
            else:
                messages.error(request, 'Данный номер телефона нельзя использовать.')

    form = UpdateProfileSellerForm(instance=request.user.profile)
    context.update({'form':form})
    return TemplateResponse(request, 'seller/requisites.html', context=context)


@login_required(login_url='/accounts/login/')
def payment(request):
    """Оплата сервиса
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Оплата сервиса',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }
    return TemplateResponse(request, 'seller/payment.html', context=context)

@login_required(login_url='/accounts/login/')
def closed_orders(request, order_number):
    """Закрывает счёт, либо показывает информацию по закрытому счёту
    """
    if request.user.profile.type != 4:
        raise PermissionDenied

    context = {
        # Заголовок текущей страницы
        'title': 'Закрыте счета',
        # отображаемое на сайте имя пользователя
        'screenname': request.user.first_name or request.user.username
    }

    if order_number < 1:
        messages.danger(request, 'Заказа с таким номером не существует.')
    else:

        if Order.objects.filter(order_number=order_number):
        
            order = Order.objects.filter(order_number=order_number).first()

            if order.status != 4:
                order.status = 4
                order.save()
                messages.success(request, f'Заказ {order_number} отмечен как закрытый.')
            context.update({'current_order': order})
            # Сумма заказа
            total_sum = 0
            for item in Order.objects.filter(order_number=order_number):
                total_sum += item.product_count * item.product.price
            context.update({'order_sum': total_sum})
        else:
            messages.danger(request, 'Заказа с таким номером не найдено.')

    # Все закрытые заказы пользователя
    # status=4
    order_ids = Order.objects.filter(status=4).values('order_number').filter(product__user=request.user).distinct()
    context.update({'closed_count': order_ids.count()})

    if order_ids.count() > 0:
        order_list = []
        for n in order_ids:
            i = n['order_number']
            order_items = Order.objects.filter(order_number=i)
            # Сумма заказа
            total_sum = 0
            for item in order_items:
                total_sum += item.product_count * item.product.price
            ordr = {
                'order_data': Order.objects.filter(order_number=i).first(),
                'order_sum': total_sum
            }
            order_list.append(ordr)
        context.update({'order_list': order_list})

    # Order.objects.filter(status=2).filter(product__user=request.user).distinct()
    

    return TemplateResponse(request, 'seller/closed_orders.html', context=context)